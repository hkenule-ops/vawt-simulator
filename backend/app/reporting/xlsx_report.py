"""
Excel workbook report generator using openpyxl. Multiple sheets: a summary
sheet, and detailed data tables (Cp-lambda curve) for further analysis in
Excel -- the point of an .xlsx export vs. a PDF/DOCX narrative is that the
underlying numbers are usable, not just readable.
"""
from __future__ import annotations
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.reporting.report_data import ReportData

HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws, n_cols: int):
    for col in range(1, n_cols + 1):
        letter = get_column_letter(col)
        max_len = max((len(str(cell.value)) for cell in ws[letter] if cell.value is not None), default=10)
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


def generate_xlsx_report(data: ReportData) -> bytes:
    wb = Workbook()

    summary = wb.active
    summary.title = "Summary"
    summary.append(["Hybrid VAWT Design Report", data.geometry.name])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append([])
    summary.append(["Parameter", "Value"])
    _style_header_row(summary, 3, 2)
    rows = [
        ("Target power (W)", data.geometry.target_power_w),
        ("Material", data.material_key),
        ("Operating wind speed (m/s)", data.wind_speed_ms),
        ("Operating TSR", data.operating_tsr),
        ("", ""),
        ("Peak Cp", round(data.peak_cp, 4)),
        ("Peak Cp at TSR", round(data.peak_cp_tsr, 3)),
        ("Structural safety factor", round(data.structural_safety_factor, 3)),
        ("Max stress (MPa)", round(data.structural_max_stress_pa / 1e6, 2)),
        ("Max deflection (mm)", round(data.structural_max_deflection_m * 1000, 3)),
        ("Spar mass (kg)", round(data.spar_mass_kg, 4)),
        ("CFRP spar mass (g)", round(data.composite_cfrp_mass_kg * 1000, 1)),
        ("GFRP spar mass (g)", round(data.composite_gfrp_mass_kg * 1000, 1)),
        ("Fatigue life (years)", data.fatigue_life_years if data.fatigue_life_years < 1e6 else "1,000,000+"),
        ("Natural frequency mode 1 (Hz)", round(data.natural_frequencies_hz[0], 2) if data.natural_frequencies_hz else None),
        ("AEP (kWh/yr)", round(data.aep_kwh, 1)),
        ("Capacity factor", round(data.capacity_factor, 4)),
        ("Total CAPEX (USD)", round(data.total_capex_usd, 2)),
        ("LCOE (USD/kWh)", round(data.lcoe_usd_per_kwh, 4)),
        ("NPV (USD)", round(data.npv_usd, 2)),
        ("IRR", round(data.irr, 4) if data.irr is not None else "N/A"),
        ("Simple payback (years)", round(data.payback_years, 2) if data.payback_years != float("inf") else "never"),
    ]
    for k, v in rows:
        summary.append([k, v])
    _autosize(summary, 2)

    cp_sheet = wb.create_sheet("Cp-Lambda Curve")
    cp_sheet.append(["Tip-Speed Ratio", "System Cp", "Total Power (W)", "Induction Factor"])
    _style_header_row(cp_sheet, 1, 4)
    for p in data.cp_lambda_points:
        cp_sheet.append([round(p.tip_speed_ratio, 3), round(p.system_cp, 4),
                          round(p.total_power_w, 1), round(p.induction_factor, 4)])
    _autosize(cp_sheet, 4)

    warnings_sheet = wb.create_sheet("Warnings")
    warnings_sheet.append(["Warning"])
    _style_header_row(warnings_sheet, 1, 1)
    for w in data.warnings:
        warnings_sheet.append([w])
    _autosize(warnings_sheet, 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
