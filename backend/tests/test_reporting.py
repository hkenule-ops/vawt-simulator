"""
Report generation tests. Two kinds of checks:
1. The assembled ReportData matches calling the underlying Stage 1-7
   functions directly (proving the report isn't a separate, potentially
   drifted code path).
2. Each generated file is structurally valid for its format (a real docx
   zip with document.xml, a real xlsx workbook openpyxl can re-read, a
   real PDF with the %PDF header, parseable CSV rows) -- not just "did it
   not crash."
"""
import io
import zipfile
import csv as csv_module
import pytest

from app.geometry.models import HybridRotorGeometry
from app.reporting.report_data import assemble_report_data
from app.reporting.docx_report import generate_docx_report
from app.reporting.xlsx_report import generate_xlsx_report
from app.reporting.pdf_report import generate_pdf_report
from app.reporting.csv_export import generate_csv_export


@pytest.fixture(scope="module")
def report_data():
    return assemble_report_data(HybridRotorGeometry())


def test_report_data_peak_cp_matches_direct_bem_call(report_data):
    from app.aero.hybrid_solver import cp_lambda_curve
    geom = HybridRotorGeometry()
    points = cp_lambda_curve(geom, geom.rated_wind_speed_ms, tsr_min=0.5, tsr_max=4.5, n_points=20)
    expected_peak = max(p.system_cp for p in points)
    assert report_data.peak_cp == pytest.approx(expected_peak, rel=1e-6)


def test_report_data_safety_factor_matches_direct_structural_call(report_data):
    from app.structural.blade_analysis import analyze_blade_structure
    geom = HybridRotorGeometry()
    struct = analyze_blade_structure(geom, "CFRP_UD", geom.rated_wind_speed_ms, 2.25,
                                      spar_width_fraction=0.5, spar_wall_thickness_m=0.003)
    assert report_data.structural_safety_factor == pytest.approx(struct.safety_factor, rel=1e-6)


def test_report_data_aep_matches_direct_economics_call(report_data):
    from app.economics.economic_analysis import analyze_economics
    geom = HybridRotorGeometry()
    econ = analyze_economics(geom, spar_mass_kg=report_data.spar_mass_kg, ply_material_key="CFRP_UD_PLY")
    assert report_data.aep_kwh == pytest.approx(econ.aep.aep_kwh, rel=1e-6)


def test_docx_report_is_valid_zip_with_document_xml(report_data):
    docx_bytes = generate_docx_report(report_data)
    zf = zipfile.ZipFile(io.BytesIO(docx_bytes))
    assert "word/document.xml" in zf.namelist()
    assert zf.testzip() is None  # no corrupted entries


def test_docx_report_contains_key_figures_as_text(report_data):
    from docx import Document
    docx_bytes = generate_docx_report(report_data)
    doc = Document(io.BytesIO(docx_bytes))
    full_text = "\n".join(p.text for p in doc.paragraphs)
    assert report_data.geometry.name in full_text
    assert "Betz" in full_text or "0.593" in full_text


def test_xlsx_report_has_expected_sheets(report_data):
    import openpyxl
    xlsx_bytes = generate_xlsx_report(report_data)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert "Summary" in wb.sheetnames
    assert "Cp-Lambda Curve" in wb.sheetnames
    assert wb["Cp-Lambda Curve"].max_row == len(report_data.cp_lambda_points) + 1  # +1 header


def test_xlsx_cp_lambda_data_matches_report_data(report_data):
    import openpyxl
    xlsx_bytes = generate_xlsx_report(report_data)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb["Cp-Lambda Curve"]
    first_data_row = [ws.cell(row=2, column=c).value for c in range(1, 5)]
    expected = report_data.cp_lambda_points[0]
    assert first_data_row[0] == pytest.approx(expected.tip_speed_ratio, rel=1e-3)
    assert first_data_row[1] == pytest.approx(expected.system_cp, rel=1e-3)


def test_pdf_report_has_valid_pdf_header(report_data):
    pdf_bytes = generate_pdf_report(report_data)
    assert pdf_bytes[:5] == b"%PDF-"


def test_pdf_report_is_reasonably_sized(report_data):
    """A near-empty or truncated PDF would be suspiciously small."""
    pdf_bytes = generate_pdf_report(report_data)
    assert len(pdf_bytes) > 2000


def test_csv_export_is_parseable_and_has_expected_rows(report_data):
    csv_bytes = generate_csv_export(report_data)
    text = csv_bytes.decode("utf-8")
    rows = list(csv_module.reader(io.StringIO(text)))
    assert any(row and row[0] == "Peak Cp" for row in rows)
    assert any(row and row[0] == "Cp-Lambda Curve" for row in rows)


def test_csv_export_cp_values_match_report_data(report_data):
    csv_bytes = generate_csv_export(report_data)
    text = csv_bytes.decode("utf-8")
    rows = list(csv_module.reader(io.StringIO(text)))
    peak_cp_row = next(row for row in rows if row and row[0] == "Peak Cp")
    assert float(peak_cp_row[1]) == pytest.approx(report_data.peak_cp, rel=1e-3)
